import docker
import sys
import os
import json
import itertools
from iteration_utilities import random_product
import random
import hashlib
from time import sleep
import shutil
import pathlib
import openpyxl
import pandas as pd
from typing import Dict

TESTING_COMPETITORS = ["humming_trader"]
MAX_NUMBER_PARAMETER_COMBINATIONS = 30 
MAX_CONCURRENT_SIMULATIONS = 2 

default_exchange_settings = {
  "Engine": {
    "MarketDataFile": "data/market_data2.csv",
    "MarketEventInterval": 0.05,
    "MarketOpenDelay": 5.0,
    "MatchEventsFile": "match_events.csv",
    "ScoreBoardFile": "score_board.csv",
    "Speed": 2.0,
    "TickInterval": 0.25
  },
  "Execution": {
    "Host": "127.0.0.1",
    "Port": 12345
  },
  "Fees": {
    "Maker": -0.0001,
    "Taker": 0.0002
  },
  "Hud": {
    "Host": "127.0.0.1",
    "Port": 12347
  },
  "Information": {
    "Type": "mmap",
    "Name": "info.dat"
  },
  "Instrument": {
    "EtfClamp": 0.002,
    "TickSize": 1.00
  },
  "Limits": {
    "ActiveOrderCountLimit": 10,
    "ActiveVolumeLimit": 200,
    "MessageFrequencyInterval": 1.0,
    "MessageFrequencyLimit": 50,
    "PositionLimit": 100
  },
  "Traders": {
  }
}

def get_next_parameter_combination(parameters_file : str):
    if not os.path.exists(parameters_file):
        print("Path doesn't exist", parameters_file)
        return {}

    with open(parameters_file, "r") as file:
        parameter_options = json.load(file)

        parameter_names = []
        permutable_parameter_values = []
        number_permutations = 1

        for parameter_name, param_options in parameter_options.items():
            parameter_names.append(parameter_name)
            permutable_parameter_values.append(param_options)
            number_permutations *= len(param_options)

        # We'll test at most 128 different combinations of the parameters
        number_permutations = min(number_permutations, MAX_NUMBER_PARAMETER_COMBINATIONS)

        permutations = random_product(*permutable_parameter_values, repeat=4 * number_permutations)

        for i in range(len(permutations))[::len(parameter_names)]:
            obj = {}
            for j in range(len(parameter_names)):
                obj[parameter_names[j]] = permutations[i + j]
            yield obj

    yield None

def read_market_file_from_trader_parameters(file):
    return json.load(open(file, "r"))["Parameters"]["MarketDataFile"]

def create_report(score_board_files, main_trader, parameters_for_each_match, report_path):
    open(report_path, "w").close()
    wb = openpyxl.Workbook()
    ws = wb.active

    # things for the colors
    my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
    my_green = openpyxl.styles.colors.Color(rgb='00CC66')
    my_grey = openpyxl.styles.colors.Color(rgb='C9C1C1')
    red_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
    green_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_green)
    grey_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_grey)

    columns = ['Team', 'BuyVolume', 'SellVolume', 'EtfPosition', 'FuturePosition',
       'EtfPrice', 'FuturePrice', 'TotalFees', 'AccountBalance',
       'ProfitOrLoss', 'Status']

    def write_match_outcome(row_start, outcome, main_trader="", market_file="", hex_code="", parameters={}) -> int:
        row_index = row_start
        
        ws.cell(row=row_index, column=1, value="Match {0}".format(hex_code))
        ws.cell(row=row_index, column=2, value=main_trader)
        ws.cell(row=row_index, column=3, value=market_file)
        ws.cell(row=row_index, column=4, value=str(parameters))
        
        column_index = 1
        row_index += 1
        for column in columns:
            ws.cell(row=row_index, column=column_index, value=column)
            column_index += 1
            
        row_index += 1
        for trader_outcome in outcome:
            column_index = 1
            for column in columns:
                if column == "ProfitOrLoss":
                    if float(trader_outcome[column]) < 0:
                        cell = ws.cell(row=row_index, column=column_index, value=trader_outcome[column])
                        cell.fill = red_fill
                    else:
                        cell = ws.cell(row=row_index, column=column_index, value=trader_outcome[column])
                        cell.fill = green_fill
                elif trader_outcome["Team"] == main_trader and column == "Team":
                    cell = ws.cell(row=row_index, column=column_index, value=trader_outcome[column])
                    cell.fill = grey_fill
                else:
                    ws.cell(row=row_index, column=column_index, value=trader_outcome[column])
                column_index += 1
            row_index += 1
            
        row_index += 2 # separation to next thing we write
        return row_index

    row_index = 1 

    score_board_file_and_pnl = []

    for i in range(len(score_board_files)):
        score_board = pd.read_csv(score_board_files[i])
        main_trader_pnl = float(score_board[score_board["Team"]==main_trader].iloc[-1]["ProfitOrLoss"])
        print("PnL", main_trader_pnl)
        score_board_file_and_pnl.append((main_trader_pnl, score_board_files[i], parameters_for_each_match[i]))

    score_board_file_and_pnl = sorted(score_board_file_and_pnl, key=lambda x: -x[0])

    i = 1
    for pnl, file, params in score_board_file_and_pnl:
        score_board = pd.read_csv(file)
        print("Storing match #{0} information:".format(i+1))
        print(score_board.tail())
        traders = list(score_board["Team"].unique()) 
        match_outcome = []
        for trader in traders:
            trader_outcome = score_board[score_board["Team"] == trader].iloc[-1]
            trader_outcome["EtfPrice"] /= 100
            trader_outcome["FuturePrice"] /= 100
            trader_outcome["TotalFees"] /= 100
            trader_outcome["AccountBalance"] /= 100
            trader_outcome["ProfitOrLoss"] /= 100
            match_outcome.append(trader_outcome) 
        match_outcome.sort(key=lambda stats: -stats["ProfitOrLoss"])
        
        settings_path = os.path.join(pathlib.Path(file).parent, "{0}.json".format(main_trader))
        market_file = read_market_file_from_trader_parameters(settings_path)
        row_index = write_match_outcome(row_index, match_outcome, main_trader=main_trader, market_file=market_file, hex_code="#"+str(i+1), parameters=params)
        i += 1

    print("Wrote {0} match reports in the excel file".format(len(score_board_files)))
    
    wb.save(report_path)  # save the workbook
    wb.close()  # close the workbook

def run_batch_of_containers(client, mounts, containers_info):
    running_containers = []

    for trader_name, hex_code, command, config in containers_info:
        print("Running container {0}.{1} testing {0} with parameters {2}".format(trader_name, hex_code, config))
        running_containers.append(client.containers.run(image="trader-test",
                                mounts=mounts,
                                name="{0}.{1}".format(trader_name, hex_code),
                                command=command,
                                detach=True))

    while len(running_containers) > 0:
        containers_to_remove = []

        for i in range(len(running_containers)):
            container = client.containers.get(running_containers[i].id)
            container.reload()

            dkg = container.logs(stream = True, follow = False)
            try:
                while True:
                    line = next(dkg).decode("utf-8")
                    print("[{0}]".format(container.name), line)
            except StopIteration:
                # print(f'Log stream ended for {container.name}')   
                pass

            status = client.containers.get(running_containers[i].id).status
            if status == "exited" or status == "removed":
                containers_to_remove.append(i)

        for index in containers_to_remove[::-1]:
            container = client.containers.get(running_containers[index].id)
            container.remove()
            running_containers.pop(index)

        sleep(1)

    print("Finished running batch of containers")


def get_mounts():
    user_working_dir = os.getcwd()

    # mount target dir needs to be absolute
    ready_trader_go_code = docker.types.Mount(target="/pyready_trader_go/ready_trader_go",
                                        source=user_working_dir+"/ready_trader_go",
                                        type="bind")

    trading_strategies = docker.types.Mount(target="/pyready_trader_go/traders",
                                      source=user_working_dir+"/traders",
                                      type="bind")

    rtg_file = docker.types.Mount(target="/pyready_trader_go/rtg.py",
                                  source=user_working_dir+"/rtg.py",
                                  type="bind")

    mounts = [ready_trader_go_code, trading_strategies, rtg_file]

    return mounts 

def run_benchmark(trader_name):
    config_file = os.path.join("traders", trader_name, trader_name + ".json")
    parameters_file = os.path.join("traders", trader_name, "testing_parameters.json") 

    default_exchange_settings["Traders"] = dict() 
    for trader in TESTING_COMPETITORS:
        default_exchange_settings["Traders"][trader] = "secret"
    default_exchange_settings["Traders"][trader_name] = "secret"

    if trader_name is None or len(trader_name) == 0 or not os.path.exists(config_file):
        print("Trader doesn't exist!")
        print("Please create a folder for the trading strategy that will be tested")
        exit(1)

    folders_before = os.listdir(os.path.join("traders", trader_name, "logs"))

    trader_config = json.load(open(config_file, "r"))

    client = docker.from_env()
    mounts = get_mounts()

    trader_code_hash = ""
    with open(os.path.join("traders", trader_name, trader_name + ".py")) as file:
        trader_code_hash = hashlib.sha256(file.read().encode()).hexdigest()

    tried_combinations = []
    generator = get_next_parameter_combination(parameters_file)
    testing_done = False
    batch_index = 1

    while not testing_done:
        next_batch_of_containers = []
        
        while not testing_done and len(next_batch_of_containers) < MAX_CONCURRENT_SIMULATIONS:
            if len(tried_combinations) >= MAX_NUMBER_PARAMETER_COMBINATIONS:
                testing_done = True
                continue

            parameters = next(generator)

            if parameters == None:
                # no more parameter combinations to try, we're done
                testing_done = True
            else:
                trader_config["Parameters"] = parameters
                trader_config["Exchange"] = default_exchange_settings
                trader_config["Exchange"]["MarketDataFile"] = parameters["MarketDataFile"] 

                hex_code = hashlib.sha256((trader_code_hash + str(parameters)).encode()).hexdigest()[:6]

                # we don't want to repeat a simulation with parameters we've already tried
                if hex_code in tried_combinations:
                    continue

                tried_combinations.append(hex_code)

                config_path = os.path.join("traders", trader_name, "{0}.{1}.json".format(trader_name, hex_code))
                json.dump(trader_config, open(config_path, "w"), indent=4)

                command = "{0}.{1}".format(trader_name, hex_code)
                for competitor_name in TESTING_COMPETITORS:
                    command += " " + competitor_name

                next_batch_of_containers.append((trader_name, hex_code, command, str(trader_config)))

        if len(next_batch_of_containers) == 0:
            assert testing_done == True
            continue
            
        run_batch_of_containers(client, mounts, next_batch_of_containers)

        if len(tried_combinations) > MAX_NUMBER_PARAMETER_COMBINATIONS:
            testing_done = True

    folders_now = os.listdir(os.path.join("traders", trader_name, "logs"))

    simulation_scores = []
    parameters_for_each_match = []

    for folder in folders_now:
        if not folder in folders_before:
            with open(os.path.join("traders", trader_name, "logs", folder, trader_name + ".json"), "r") as file:
                parameters_for_each_match.append(json.load(file)["Parameters"])
                simulation_scores.append(os.path.join("traders", trader_name, "logs", folder, "score_board.csv"))

    create_report(simulation_scores, trader_name, parameters_for_each_match, "benchmark_reports/benchmark_report.xlsx")



if __name__ == "__main__":
    trader_name = sys.argv[1]
    run_benchmark(trader_name)
    